# coding: utf-8
# CS FutureMobility Tool
# See full license in LICENSE.txt.

import numpy as np
import pandas as pd
#import openmatrix as omx
from IPython.display import display
from openpyxl import load_workbook,Workbook
from time import strftime
import os.path
import mode_choice.model_defs as md
import mode_choice.matrix_utils as mtx
import config
''' Utilities to summarize the outputs of Mode Choice '''

    
def display_mode_share(mc_obj):
    '''
    This displays a mode share summary by market segment (with / without vehicle, peak / off-peak) on the IPython notebook.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    '''
    # display mode share tables
    avg_trips_by_mode = pd.DataFrame(None)
    for purpose in ['HBW','HBO', 'NHB', 'HBSc1', 'HBSc2', 'HBSc3']:
        avg_trips_by_mode = avg_trips_by_mode.add(pd.DataFrame({pv:{mode:(mc_obj.table_container.get_table(purpose)[pv][mode].sum()) for mode in mc_obj.table_container.get_table(purpose)[pv]} for pv in ['0_PK','1_PK','0_OP','1_OP']}).T,
            fill_value = 0)
    avg_mode_share = avg_trips_by_mode.divide(avg_trips_by_mode.sum(1),axis = 0)

    display(avg_mode_share.style.format("{:.2%}"))
    
    
def write_boston_neighbortown_mode_share_to_excel(mc_obj):
    '''
    Writes mode share summary by purpose and market segment to an Excel workbook.
    Applies only to trips to/from Boston
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param out_excel_fn: output Excel filename, by default in the output path defined in config.py  
    '''
    
    out_excel_fn = mc_obj.config.out_path + "mode_share_bosNB_{0}.xlsx".format(strftime("%Y%m%d"))

    # check if file exists.
    if os.path.isfile(out_excel_fn):
        book = load_workbook(out_excel_fn)
    else: 
        book = Workbook()
        book.save(out_excel_fn)
    writer = pd.ExcelWriter(out_excel_fn,engine = 'openpyxl')
    writer.book = book
    
    for purp in md.purposes:
        mode_share = pd.DataFrame(columns = md.peak_veh)
        trip_table = mc_obj.table_container.get_table(purp)
        for pv in md.peak_veh:
            for mode in trip_table[pv].keys():
            #study area zones might not start at zone 0 and could have discontinous TAZ IDs
                
                trip_table_o = mtx.OD_slice(trip_table[pv][mode], O_slice = md.taz['BOSTON'], D_slice = md.taz['BOS_AND_NEI'])
                trip_table_d = mtx.OD_slice(trip_table[pv][mode], O_slice = md.taz['BOS_AND_NEI'], D_slice = md.taz['BOSTON'])
                trip_table_b = mtx.OD_slice(trip_table[pv][mode], O_slice = md.taz['BOSTON'], D_slice = md.taz['BOSTON'])
                
                trip_table_bos = trip_table_o + trip_table_d - trip_table_b
                mode_share.loc[mode,pv] = trip_table_bos.sum()

        mode_share['Total'] = mode_share.sum(1)
        mode_share['Share'] = mode_share['Total'] / mode_share['Total'].sum()
        
        if purp in book.sheetnames: # if sheetname exists, delete
            book.remove(book[purp])
            writer.save()

        mode_share.to_excel(writer, sheet_name = purp)
    
        writer.save()

def write_study_area_mode_share_to_excel(mc_obj, out_excel_fn = None):
    '''
    Writes mode share summary by purpose and market segment to an Excel workbook.
    Applies only to trips to/from study area
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param out_excel_fn: output Excel filename, by default in the output path defined in config.py  
    '''
    if out_excel_fn is None:
        out_excel_fn = mc_obj.config.out_path + "mode_share_study_area_{0}.xlsx".format(strftime("%Y%m%d"))
    
    # check if file exists.
    if os.path.isfile(out_excel_fn):
        book = load_workbook(out_excel_fn)
    else: 
        book = Workbook()
        book.save(out_excel_fn)
    writer = pd.ExcelWriter(out_excel_fn,engine = 'openpyxl')
    writer.book = book
    
    for purp in md.purposes:
        mode_share = pd.DataFrame(columns = md.peak_veh)
        trip_table = mc_obj.table_container.get_table(purp)
        for pv in md.peak_veh:
            for mode in trip_table[pv].keys():
            
                trip_table_o = mtx.OD_slice(trip_table[pv][mode], O_slice = md.study_area)
                trip_table_d = mtx.OD_slice(trip_table[pv][mode], D_slice = md.study_area)
                trip_table_ii = mtx.OD_slice(trip_table[pv][mode], O_slice = md.study_area, D_slice = md.study_area)
            
                trip_table_sa = trip_table_o + trip_table_d - trip_table_ii
                mode_share.loc[mode,pv] = trip_table_sa.sum()

        mode_share['Total'] = mode_share.sum(1)
        mode_share['Share'] = mode_share['Total'] / mode_share['Total'].sum()
        
        if purp in book.sheetnames: # if sheetname exists, delete
            book.remove(book[purp])
            writer.save()

        mode_share.to_excel(writer, sheet_name = purp)
    
        writer.save()
    
def write_mode_share_to_excel(mc_obj,purpose, out_excel_fn = None):
    '''
    Writes mode share summary by purpose and market segment to an Excel workbook.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param purpose: can be a single purpose or 'all', in which case the Excel workbook has six sheets, one for each purpose.
    :param out_excel_fn: output Excel filename, by default in the output path defined in config.py  
    '''
    if out_excel_fn is None:
        out_excel_fn = mc_obj.config.out_path + "MC_mode_share_{0}_{1}.xlsx".format(purpose, strftime("%Y%m%d"))
    if purpose == 'all':
        # check if file exists.
        if os.path.isfile(out_excel_fn):
            book = load_workbook(out_excel_fn)
        else: 
            book = Workbook()
            book.save(out_excel_fn)
        writer = pd.ExcelWriter(out_excel_fn,engine = 'openpyxl')
        writer.book = book
        
        for purp in md.purposes:
            trip_table = mc_obj.table_container.get_table(purp)
            mode_share = pd.DataFrame(columns = md.peak_veh)
            for pv in md.peak_veh:
                for mode in trip_table[pv].keys():
                    mode_share.loc[mode,pv] = trip_table[pv][mode].sum()
            
            mode_share['Total'] = mode_share.sum(1)
            mode_share['Share'] = mode_share['Total'] / mode_share['Total'].sum()
            
            if purp in book.sheetnames: # if sheetname exists, delete
                book.remove(book[purp])
                writer.save()

            mode_share.to_excel(writer, sheet_name = purp)
        
            writer.save()
        
    elif purpose in md.purposes:
        # check if file exists.
        if os.path.isfile(out_excel_fn):
            book = load_workbook(out_excel_fn)
        else: 
            book = Workbook()
            book.save(out_excel_fn)
        writer = pd.ExcelWriter(out_excel_fn,engine = 'openpyxl')
        writer.book = book
        
        mode_share = pd.DataFrame(columns = md.peak_veh)
        for pv in md.peak_veh:
            for mode in mc_obj.trips_by_mode[pv].keys():
                mode_share.loc[mode,pv] = mc_obj.trips_by_mode[pv][mode].sum()
        
        mode_share['Total'] = mode_share.sum(1)
        mode_share['Share'] = mode_share['Total'] / mode_share['Total'].sum()
        
        if purpose in book.sheetnames: # if sheetname exists, delete
            book.remove(book[purpose])
            writer.save()

        mode_share.to_excel(writer, sheet_name = purpose)
    
        writer.save()
        
def __mt_prod_attr_nhood(mc_obj, trip_table, skim): # miles traveled. For VMT and PMT, by neighborhood
    # sum prodct of trip_table - skims
    mt_total = trip_table * skim['Length (Skim)'] 
    
    # calculate marginals
    prod = pd.DataFrame(np.sum(mt_total,axis = 1)/2, columns = ['Production'])
    attr = pd.DataFrame(np.sum(mt_total,axis = 0) / 2, columns = ['Attraction'])
    
    towns = mc_obj.taz.sort_values(md.taz_ID_field).iloc[0:md.max_zone]
    mt_taz = pd.concat([towns[[md.taz_ID_field,'BOSTON_NB']],prod,attr],axis = 1,join = 'inner')
    mt_taz.index.names=['Boston Neighborhood']
    return mt_taz.groupby(['BOSTON_NB']).sum()[['Production','Attraction']].reset_index()
    
def __trip_prod_attr_nhood(mc_obj, trip_table): 
    mt_total = trip_table
    
    # calculate marginals
    prod = pd.DataFrame(np.sum(mt_total,axis = 1), columns = ['Production'])
    attr = pd.DataFrame(np.sum(mt_total,axis = 0), columns = ['Attraction'])
    
    towns = mc_obj.taz.sort_values(md.taz_ID_field).iloc[0:md.max_zone]
    mt_taz = pd.concat([towns[[md.taz_ID_field,'BOSTON_NB']],prod,attr],axis = 1,join = 'inner')
    mt_taz.index.names=['Boston Neighborhood']
    return mt_taz.groupby(['BOSTON_NB']).sum()[['Production','Attraction']].reset_index()
    
def sm_vmt_by_neighborhood(mc_obj, out_fn = None, by = None, sm_mode = 'SM_RA'):
    '''
    Summarizes VMT production and attraction by the 26 Boston neighborhoods for Shared Mobility Modes.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param out_fn: output csv filename; if None specified, in the output path defined in config.py  
    :param by: grouping used for the summary; if None specified, only aggregate production and attraction will be provided.
    '''
    
    if out_fn is None and by is None:
        out_fn = mc_obj.config.out_path + sm_mode + f'_vmt_by_neighborhood.csv'
    elif out_fn is None and by:
        out_fn = mc_obj.config.out_path + sm_mode + f'_vmt_by_neighborhood_by_{by}.csv'
    skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
    if by in ['peak','veh_own','purpose'] == False:
        print('Only supports VMT by neighborhood, peak / vehicle ownership, purpose.')
        return
        
    else:
        vmt_master_table = pd.DataFrame(columns = ['Production','Attraction','peak','veh_own','purpose'])
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        auto_trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][sm_mode] / md.AO_dict[sm_mode]
                        vmt_table = __mt_prod_attr_nhood(mc_obj,auto_trip_table,skim_dict[peak])
                        vmt_table['peak'] = peak
                        vmt_table['veh_own'] = veh_own
                        vmt_table['purpose'] = purpose
                        vmt_master_table = vmt_master_table.append(vmt_table, sort = True)
        
        if by == None:
            vmt_summary = vmt_master_table.groupby('BOSTON_NB').sum()
        
        elif by == 'peak':
            vmt_summary = pd.concat([
            vmt_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak] for peak in ['PK','OP']], axis = 1, keys = ['PK','OP'])
        
        elif by == 'veh_own':
            vmt_summary = pd.concat([
            vmt_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own] for veh_own in ['0','1']], axis = 1, keys = ['No car', 'With car']
            )
            
        elif by == 'purpose':
            vmt_summary = pd.concat([
            vmt_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose] for purpose in vmt_master_table.purpose.unique()],axis = 1, keys= vmt_master_table.purpose.unique())

        vmt_summary.to_csv(out_fn)

def vmt_by_neighborhood(mc_obj, out_fn = None, by = None):
    '''
    Summarizes VMT production and attraction by the 26 Boston neighborhoods.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param out_fn: output csv filename; if None specified, in the output path defined in config.py  
    :param by: grouping used for the summary; if None specified, only aggregate production and attraction will be provided.
    '''
    
    if out_fn is None and by is None:
        out_fn = mc_obj.config.out_path + f'vmt_by_neighborhood.csv'
    elif out_fn is None and by:
        out_fn = mc_obj.config.out_path + f'vmt_by_neighborhood_by_{by}.csv'
    skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
    if by in ['peak','veh_own','purpose'] == False:
        print('Only supports VMT by neighborhood, peak / vehicle ownership, purpose.')
        return
        
    else:
        vmt_master_table = pd.DataFrame(columns = ['Production','Attraction','peak','veh_own','purpose'])
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        
                        auto_trip_table = sum([
                            mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] / md.AO_dict[mode]
                             for mode in ['DA','SR2','SR3+','SM_RA','SM_SH'] if mode in drive_modes])
                        
                        
                        
                        vmt_table = __mt_prod_attr_nhood(mc_obj,auto_trip_table,skim_dict[peak])
                        vmt_table['peak'] = peak
                        vmt_table['veh_own'] = veh_own
                        vmt_table['purpose'] = purpose
                        vmt_master_table = vmt_master_table.append(vmt_table, sort = True)
        
        if by == None:
            vmt_summary = vmt_master_table.groupby('BOSTON_NB').sum()
        
        elif by == 'peak':
            vmt_summary = pd.concat([
            vmt_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak] for peak in ['PK','OP']], axis = 1, keys = ['PK','OP'])
        
        elif by == 'veh_own':
            vmt_summary = pd.concat([
            vmt_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own] for veh_own in ['0','1']], axis = 1, keys = ['No car', 'With car']
            )
            
        elif by == 'purpose':
            vmt_summary = pd.concat([
            vmt_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose] for purpose in vmt_master_table.purpose.unique()],axis = 1, keys= vmt_master_table.purpose.unique())

        vmt_summary.to_csv(out_fn)

def pmt_by_neighborhood(mc_obj, out_fn = None, by = None):
    '''
    Summarizes PMT production and attraction by the 26 Boston neighborhoods.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param out_fn: output csv filename; if None specified, in the output path defined in config.py  
    :param by: grouping used for the summary; if None specified, only aggregate production and attraction will be provided.
    '''
    
    if out_fn is None and by is None:
        out_fn = mc_obj.config.out_path + f'pmt_by_neighborhood.csv'
    elif out_fn is None and by:
        out_fn = mc_obj.config.out_path + f'pmt_by_neighborhood_by_{by}.csv'
        
    skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
    
    if by in ['peak','veh_own','purpose'] == False:
        print('Only supports PMT by neighborhood, peak / vehicle ownership, purpose.')
        return
        
    else:
        pmt_master_table = pd.DataFrame(columns = ['Production','Attraction','peak','veh_own','purpose'])
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        person_trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] for mode in md.modes if mode in drive_modes])
                        
                        pmt_table = __mt_prod_attr_nhood(mc_obj,person_trip_table,skim_dict[peak])
                        pmt_table['peak'] = peak
                        pmt_table['veh_own'] = veh_own
                        pmt_table['purpose'] = purpose
                        pmt_master_table = pmt_master_table.append(pmt_table, sort = True)
        
        if by == None:
            pmt_summary = pmt_master_table.groupby('BOSTON_NB').sum()
        
        elif by == 'peak':
            pmt_summary = pd.concat([
            pmt_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak] for peak in ['PK','OP']], axis = 1, keys = ['PK','OP'])
        
        elif by == 'veh_own':
            pmt_summary = pd.concat([
            pmt_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own] for veh_own in ['0','1']], axis = 1, keys = ['No car', 'With car']
            )
            
        elif by == 'purpose':
            pmt_summary = pd.concat([
            pmt_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose] for purpose in pmt_master_table.purpose.unique()],axis = 1, keys= pmt_master_table.purpose.unique())

        pmt_summary.to_csv(out_fn)


def act_pmt_by_neighborhood(mc_obj, out_fn = None, by = None):
    '''
    Summarizes PMT production and attraction by the 26 Boston neighborhoods for active modes.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param out_fn: output csv filename; if None specified, in the output path defined in config.py  
    :param by: grouping used for the summary; if None specified, only aggregate production and attraction will be provided.
    '''
    
    if out_fn is None and by is None:
        out_fn = mc_obj.config.out_path + f'act_pmt_by_neighborhood.csv'
    elif out_fn is None and by:
        out_fn = mc_obj.config.out_path + f'act_pmt_by_neighborhood_by_{by}.csv'
        
    skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
    
    if by in ['peak','veh_own','purpose'] == False:
        print('Only supports PMT by neighborhood, peak / vehicle ownership, purpose.')
        return
        
    else:
        pmt_master_table = pd.DataFrame(columns = ['Production','Attraction','peak','veh_own','purpose'])
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        person_trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] for mode in ['Walk','Bike'] if mode in drive_modes])
                        
                        pmt_table = __mt_prod_attr_nhood(mc_obj,person_trip_table,skim_dict[peak])
                        pmt_table['peak'] = peak
                        pmt_table['veh_own'] = veh_own
                        pmt_table['purpose'] = purpose
                        pmt_master_table = pmt_master_table.append(pmt_table, sort = True)
        
        if by == None:
            pmt_summary = pmt_master_table.groupby('BOSTON_NB').sum()
        
        elif by == 'peak':
            pmt_summary = pd.concat([
            pmt_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak] for peak in ['PK','OP']], axis = 1, keys = ['PK','OP'])
        
        elif by == 'veh_own':
            pmt_summary = pd.concat([
            pmt_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own] for veh_own in ['0','1']], axis = 1, keys = ['No car', 'With car']
            )
            
        elif by == 'purpose':
            pmt_summary = pd.concat([
            pmt_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose] for purpose in pmt_master_table.purpose.unique()],axis = 1, keys= pmt_master_table.purpose.unique())

        pmt_summary.to_csv(out_fn)
        
def sm_trips_by_neighborhood(mc_obj, out_fn = None, by = None, sm_mode = 'SM_RA'):
    '''
    Summarizes PMT production and attraction by the 26 Boston neighborhoods for Shared Mobility Modes.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param out_fn: output csv filename; if None specified, in the output path defined in config.py  
    :param by: grouping used for the summary; if None specified, only aggregate production and attraction will be provided.
    :param sm_mode: Smart Mobility Mode name
    '''
    
    if out_fn is None and by is None:
        out_fn = mc_obj.config.out_path + sm_mode + f'_trips_by_neighborhood.csv'
    elif out_fn is None and by:
        out_fn = mc_obj.config.out_path + sm_mode + f'_trips_by_neighborhood_by_{by}.csv'
    
    if by in ['peak','veh_own','purpose'] == False:
        print('Only supports Trips by neighborhood, peak / vehicle ownership, purpose.')
        return
        
    else:
        trp_master_table = pd.DataFrame(columns = ['Production','Attraction','peak','veh_own','purpose'])
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        person_trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][sm_mode]
                        
                        trp_table = __trip_prod_attr_nhood(mc_obj,person_trip_table)
                        trp_table['peak'] = peak
                        trp_table['veh_own'] = veh_own
                        trp_table['purpose'] = purpose
                        trp_master_table = trp_master_table.append(trp_table, sort = True)
        
        if by == None:
            trp_summary = trp_master_table.groupby('BOSTON_NB').sum()
        
        elif by == 'peak':
            trp_summary = pd.concat([
            trp_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak] for peak in ['PK','OP']], axis = 1, keys = ['PK','OP'])
        
        elif by == 'veh_own':
            trp_summary = pd.concat([
            trp_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own] for veh_own in ['0','1']], axis = 1, keys = ['No car', 'With car']
            )
            
        elif by == 'purpose':
            trp_summary = pd.concat([
            trp_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose] for purpose in trp_master_table.purpose.unique()],axis = 1, keys= trp_master_table.purpose.unique())

        trp_summary.to_csv(out_fn)

def trips_by_neighborhood(mc_obj, out_fn = None, by = None):
    '''
    Summarizes PMT production and attraction by the 26 Boston neighborhoods.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param out_fn: output csv filename; if None specified, in the output path defined in config.py  
    :param by: grouping used for the summary; if None specified, only aggregate production and attraction will be provided.
    '''
    
    if out_fn is None and by is None:
        out_fn = mc_obj.config.out_path + f'trips_by_neighborhood.csv'
    elif out_fn is None and by:
        out_fn = mc_obj.config.out_path + f'trips_by_neighborhood_by_{by}.csv'

    if by in ['peak','veh_own','purpose'] == False:
        print('Only supports Trips by neighborhood, peak / vehicle ownership, purpose.')
        return
        
    else:
        trp_master_table = pd.DataFrame(columns = ['Production','Attraction','peak','veh_own','purpose'])
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        person_trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] for mode in md.modes if mode in drive_modes])
                        
                        trp_table = __trip_prod_attr_nhood(mc_obj,person_trip_table)
                        trp_table['peak'] = peak
                        trp_table['veh_own'] = veh_own
                        trp_table['purpose'] = purpose
                        trp_master_table = trp_master_table.append(trp_table, sort = True)
        
        if by == None:
            trp_summary = trp_master_table.groupby('BOSTON_NB').sum()
        
        elif by == 'peak':
            trp_summary = pd.concat([
            trp_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak] for peak in ['PK','OP']], axis = 1, keys = ['PK','OP'])
        
        elif by == 'veh_own':
            trp_summary = pd.concat([
            trp_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own] for veh_own in ['0','1']], axis = 1, keys = ['No car', 'With car']
            )
            
        elif by == 'purpose':
            trp_summary = pd.concat([
            trp_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose] for purpose in trp_master_table.purpose.unique()],axis = 1, keys= trp_master_table.purpose.unique())

        trp_summary.to_csv(out_fn)
def mode_share_by_neighborhood(mc_obj, out_fn = None, by = None):
    '''
    Summarizes mode share as the average of trips to/from the 26 Boston neighborhoods, in three categories - drive, non-motorized and transit.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param out_fn: output csv filename; if None specified, in the output path defined in config.py  
    :param by: grouping used for the summary
    '''
    if out_fn is None and by is None:
        out_fn = mc_obj.config.out_path + f'mode_share_by_neighborhood.csv'
    elif out_fn is None and by:
        out_fn = mc_obj.config.out_path + f'mode_share_by_neighborhood_by_{by}.csv'

    if by in ['peak','veh_own','purpose'] == False:
        print('Only supports mode share by neighborhood, peak / vehicle ownership, purpose.')
        return

    else:
        
        share_master_table = pd.DataFrame(columns = ['drive','non-motorized','transit','peak','veh_own','purpose'])

        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        share_table = pd.DataFrame(index = range(0,md.max_zone),columns = ['drive','non-motorized','transit','smart mobility']).fillna(0)
                        for mode in mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}']:
                        
                            trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode]
                            category = md.mode_categories[mode]
                            
                            share_table[category] += (trip_table.sum(axis = 1)+trip_table.sum(axis = 0))/2
                        
                        towns = mc_obj.taz.sort_values(md.taz_ID_field).iloc[0:md.max_zone]
                        trips = pd.concat([towns[[md.taz_ID_field,'BOSTON_NB']],share_table],axis = 1,join = 'inner').groupby(['BOSTON_NB']).sum().drop([md.taz_ID_field],axis = 1)
                        trips['peak'] = peak
                        trips['veh_own'] = veh_own
                        trips['purpose'] = purpose
                        
                        share_master_table = share_master_table.append(trips.reset_index(), sort = True)
        
        if by == None:
            trip_summary = share_master_table.groupby('BOSTON_NB').sum()
            share_summary = trip_summary.divide(trip_summary.sum(axis = 1),axis = 0)
        
        elif by == 'peak':
            share_summary = pd.concat([
            share_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak].divide(
            share_master_table.groupby(['peak','BOSTON_NB']).sum().loc[peak].sum(axis=1),axis = 0)
            for peak in ['PK','OP']
            ], axis = 1, keys = ['PK','OP'])
        
        elif by == 'veh_own':
            share_summary = pd.concat([
            share_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own].divide(
            share_master_table.groupby(['veh_own','BOSTON_NB']).sum().loc[veh_own].sum(axis=1),axis = 0)
            for veh_own in ['0','1']
            ], axis = 1, keys = ['No car', 'With car'])
            
        elif by == 'purpose':
            share_summary = pd.concat([
            share_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose].divide(
            share_master_table.groupby(['purpose','BOSTON_NB']).sum().loc[purpose].sum(axis=1),axis = 0)
            for purpose in share_master_table.purpose.unique()
            ],axis = 1, keys= share_master_table.purpose.unique())
        
        
        share_summary.to_csv(out_fn)
    
# Seaport method
def mode_share_by_subarea(mc_obj, out_fn = None, by = None):
    '''
    Summarizes mode share as the average of trips to/from the 7 Seaport sub-areas, in three categories - drive, non-motorized and transit.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param out_fn: output csv filename; if None specified, in the output path defined in config.py  
    :param by: grouping used for the summary
    '''
    if out_fn is None and by is None:
        out_fn = mc_obj.config.out_path + f'mode_share_by_subarea.csv'
    elif out_fn is None and by:
        out_fn = mc_obj.config.out_path + f'mode_share_by_subarea_by_{by}.csv'

    if by in ['peak','veh_own','purpose'] == False:
        print('Only supports mode share by subarea, peak / vehicle ownership, purpose.')
        return

    else:
        
        share_master_table = pd.DataFrame(columns = ['drive','non-motorized','transit','peak','veh_own','purpose'])

        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        share_table = pd.DataFrame(index = range(0,md.max_zone),columns = ['drive','non-motorized','transit','smart mobility']).fillna(0)
                        for mode in mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}']:
                        
                            trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode]
                            category = md.mode_categories[mode]
                            
                            share_table[category] += (trip_table.sum(axis = 1)+trip_table.sum(axis = 0))/2
                        
                        towns = mc_obj.taz.sort_values(md.taz_ID_field).iloc[0:md.max_zone]
                        towns['REPORT_AREA'] = towns['REPORT_AREA'][towns['REPORT_AREA'].isin(['South Station', 'Seaport Blvd', 'Design Center',
       'Southeast Seaport', 'BCEC', 'Fort Point', 'Broadway'])]
                        trips = pd.concat([towns[[md.taz_ID_field,'REPORT_AREA']],share_table],axis = 1,join = 'inner').groupby(['REPORT_AREA']).sum().drop([md.taz_ID_field],axis = 1)
                        trips['peak'] = peak
                        trips['veh_own'] = veh_own
                        trips['purpose'] = purpose
                        
                        share_master_table = share_master_table.append(trips.reset_index(), sort = True)
        
        if by == None:
            trip_summary = share_master_table.groupby('REPORT_AREA').sum()
            share_summary = trip_summary.divide(trip_summary.sum(axis = 1),axis = 0)
        
        elif by == 'peak':
            share_summary = pd.concat([
            share_master_table.groupby(['peak','REPORT_AREA']).sum().loc[peak].divide(
            share_master_table.groupby(['peak','REPORT_AREA']).sum().loc[peak].sum(axis=1),axis = 0)
            for peak in ['PK','OP']
            ], axis = 1, keys = ['PK','OP'])
        
        elif by == 'veh_own':
            share_summary = pd.concat([
            share_master_table.groupby(['veh_own','REPORT_AREA']).sum().loc[veh_own].divide(
            share_master_table.groupby(['veh_own','REPORT_AREA']).sum().loc[veh_own].sum(axis=1),axis = 0)
            for veh_own in ['0','1']
            ], axis = 1, keys = ['No car', 'With car'])
            
        elif by == 'purpose':
            share_summary = pd.concat([
            share_master_table.groupby(['purpose','REPORT_AREA']).sum().loc[purpose].divide(
            share_master_table.groupby(['purpose','REPORT_AREA']).sum().loc[purpose].sum(axis=1),axis = 0)
            for purpose in share_master_table.purpose.unique()
            ],axis = 1, keys= share_master_table.purpose.unique())
        
        
        share_summary.to_csv(out_fn)



def __sm_compute_summary_by_subregion(mc_obj,metric = 'VMT',subregion = 'neighboring', sm_mode='SM_RA'):
    ''' Computing function used by write_summary_by_subregion(), does not produce outputs'''

    if metric.lower() not in ('vmt','pmt','mode share','trip', 'pmt_act'):
        print('Only supports trip, VMT, PMT and mode share calculations.')
        return
    if subregion.lower() not in ('boston','neighboring','i93','i495','region'):
        print('Only supports within boston, "neighboring" for towns neighboring Boston, I93, I495 or Region.')
        return

    subregion_dict = {'boston':'BOSTON','neighboring':'BOS_AND_NEI','i93':'in_i95i93','i495':'in_i495'}
    
            
    if metric.lower() == 'vmt':
        skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
        vmt_table = np.zeros((md.max_zone,md.max_zone))
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][sm_mode] / md.AO_dict[sm_mode]
                        vmt_table += trip_table * skim_dict[peak]['Length (Skim)']
                        
        if subregion.lower() in subregion_dict:
            field = subregion_dict[subregion.lower()]
            boston_o_auto_vmt = mtx.OD_slice(vmt_table,O_slice = md.taz['BOSTON'], D_slice = md.taz[field]== True)
            boston_d_auto_vmt = mtx.OD_slice(vmt_table,md.taz[field]== True,D_slice = md.taz['BOSTON'])
            #boston_o_auto_vmt = vmt_table[md.taz['BOSTON'],:][:, md.taz[field]== True]
            #boston_d_auto_vmt = vmt_table[md.taz[field]== True,:][:,md.taz['BOSTON']]
            town_definition = md.taz[md.taz[field]== True]
        
        elif subregion.lower() == 'region':
            boston_o_auto_vmt = mtx.OD_slice(vmt_table,O_slice = md.taz['BOSTON'])
            boston_d_auto_vmt = mtx.OD_slice(vmt_table,D_slice = md.taz['BOSTON'])
            #boston_o_auto_vmt = vmt_table[md.taz['BOSTON'],:]
            #boston_d_auto_vmt = vmt_table[:][:,md.taz['BOSTON']]
            town_definition = md.taz   
        
        zone_vmt_daily_o = pd.DataFrame(np.sum(boston_o_auto_vmt,axis=1)/2 ,columns=["VMT"])
        zone_vmt_daily_d = pd.DataFrame(np.sum(boston_d_auto_vmt,axis=0)/2 ,columns=["VMT"])

        town_vmt_o=pd.concat([town_definition,zone_vmt_daily_o],axis=1,join='inner')
        town_vmt_d=pd.concat([town_definition,zone_vmt_daily_d],axis=1,join='inner')

        vmt_sum_o = town_vmt_o[town_vmt_o['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['VMT']
        vmt_sum_d = town_vmt_d[town_vmt_d['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['VMT']

        subregion_vmt = (vmt_sum_o + vmt_sum_d).values[0]
        
        return subregion_vmt
    
    elif metric.lower() == 'trip':
        skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
        tripsum_table = np.zeros((md.max_zone,md.max_zone))
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][sm_mode]
                        tripsum_table += trip_table
                        
        if subregion.lower() in subregion_dict:
            field = subregion_dict[subregion.lower()]
            boston_o_trip = mtx.OD_slice(tripsum_table, O_slice = md.taz['BOSTON'],D_slice = md.taz[field]== True)
            boston_d_trip = mtx.OD_slice(tripsum_table, O_slice = md.taz[field]== True, D_slice = md.taz['BOSTON'])
            #boston_o_trip = tripsum_table[md.taz['BOSTON'],:][:, md.taz[field]== True]
            #boston_d_trip = tripsum_table[md.taz[field]== True,:][:,md.taz['BOSTON']]
            town_definition = md.taz[md.taz[field]== True]
        
        elif subregion.lower() == 'region':
            boston_o_trip = mtx.OD_slice(tripsum_table, O_slice = md.taz['BOSTON'])
            boston_d_trip = mtx.OD_slice(tripsum_table, D_slice = md.taz['BOSTON'])
            #boston_o_trip = tripsum_table[md.taz['BOSTON'],:]
            #boston_d_trip = tripsum_table[:][:,md.taz['BOSTON']]
            town_definition = md.taz   
        
        zone_daily_o = pd.DataFrame(np.sum(boston_o_trip,axis=1) ,columns=["trips"])
        zone_daily_d = pd.DataFrame(np.sum(boston_d_trip,axis=0) ,columns=["trips"])

        town_o=pd.concat([town_definition,zone_daily_o],axis=1,join='inner')
        town_d=pd.concat([town_definition,zone_daily_d],axis=1,join='inner')

        sum_o = town_o[town_o['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['trips']
        sum_d = town_d[town_d['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['trips']

        subregion_trip = (sum_o + sum_d).values[0]
        
        return subregion_trip
        
def __compute_metric_by_zone(mc_obj,metric = 'VMT'):
    ''' Computing function used by write_summary_by_subregion(), does not produce outputs'''
            
    if metric.lower() == 'vmt':
        skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
        vmt_table = np.zeros((md.max_zone,md.max_zone))
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] / md.AO_dict[mode] for mode in md.auto_modes if mode in drive_modes])
                        vmt_table += trip_table * skim_dict[peak]['Length (Skim)']
                        
        boston_o_auto_vmt = mtx.OD_slice(vmt_table, O_slice = md.taz['BOSTON'])
        boston_d_auto_vmt = mtx.OD_slice(vmt_table,D_slice = md.taz['BOSTON'])
        
        #boston_o_auto_vmt = vmt_table[md.taz['BOSTON'],:]
        #boston_d_auto_vmt = vmt_table[:][:,md.taz['BOSTON']]
        town_definition = md.taz   
        
        zone_vmt_daily_o = pd.DataFrame(np.sum(boston_o_auto_vmt,axis=0)/2 ,columns=["VMT"])
        zone_vmt_daily_d = pd.DataFrame(np.sum(boston_d_auto_vmt,axis=1)/2 ,columns=["VMT"])
        
        town_vmt_o=pd.concat([town_definition,zone_vmt_daily_o],axis=1,join='inner')
        town_vmt_d=pd.concat([town_definition,zone_vmt_daily_d],axis=1,join='inner')
        
        town_vmt = town_vmt_o.groupby(['TOWN']).sum()['VMT'] + town_vmt_d.groupby(['TOWN']).sum()['VMT']
        
        return town_vmt
        
    elif metric.lower() == 'pmt':
        skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
        pmt_table = np.zeros((md.max_zone,md.max_zone))
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        
                        trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                        for mode in md.modes if mode in drive_modes])
                        
                        pmt_table += trip_table * skim_dict[peak]['Length (Skim)']
        boston_o_auto_pmt = mtx.OD_slice(pmt_table, O_slice = md.taz['BOSTON'])
        boston_d_auto_pmt = mtx.OD_slice(pmt_table, D_slice = md.taz['BOSTON'])
        #boston_o_auto_pmt = pmt_table[md.taz['BOSTON'],:]
        #boston_d_auto_pmt = pmt_table[:][:,md.taz['BOSTON']]
        town_definition = md.taz   
        
        zone_pmt_daily_o = pd.DataFrame(np.sum(boston_o_auto_pmt,axis=0)/2 ,columns=["VMT"])
        zone_pmt_daily_d = pd.DataFrame(np.sum(boston_d_auto_pmt,axis=1)/2 ,columns=["VMT"])
        
        town_pmt_o=pd.concat([town_definition,zone_pmt_daily_o],axis=1,join='inner')
        town_pmt_d=pd.concat([town_definition,zone_pmt_daily_d],axis=1,join='inner')
        
        town_pmt = town_pmt_o.groupby(['TOWN']).sum()['VMT'] + town_pmt_d.groupby(['TOWN']).sum()['VMT']
        
        return town_pmt
    
    elif metric.lower() == 'pmt_act':
        skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
        
        
        pmt_table = np.zeros((md.max_zone,md.max_zone))
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        
                        trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                        for mode in ['Walk','Bike'] if mode in drive_modes])
                        
                        pmt_table += trip_table * skim_dict[peak]['Length (Skim)']
        
        boston_o_auto_pmt = mtx.OD_slice(pmt_table, O_slice = md.taz['BOSTON'])
        boston_d_auto_pmt = mtx.OD_slice(pmt_table, D_slice = md.taz['BOSTON'])
        #boston_o_auto_pmt = pmt_table[taz['BOSTON'],:]
        #boston_d_auto_pmt = pmt_table[:][:,taz['BOSTON']]
        town_definition = md.taz   
        
        zone_pmt_daily_o = pd.DataFrame(np.sum(boston_o_auto_pmt,axis=0)/2 ,columns=["VMT"])
        zone_pmt_daily_d = pd.DataFrame(np.sum(boston_d_auto_pmt,axis=1)/2 ,columns=["VMT"])
        
        town_pmt_o=pd.concat([town_definition,zone_pmt_daily_o],axis=1,join='inner')
        town_pmt_d=pd.concat([town_definition,zone_pmt_daily_d],axis=1,join='inner')
        
        town_pmt = town_pmt_o.groupby(['TOWN']).sum()['VMT'] + town_pmt_d.groupby(['TOWN']).sum()['VMT']
        
        return town_pmt
        
def __compute_summary_by_subregion(mc_obj,metric = 'VMT',subregion = 'neighboring'):
    ''' Computing function used by write_summary_by_subregion(), does not produce outputs'''


    if metric.lower() not in ('vmt','pmt','mode share','trip', 'pmt_act'):
        print('Only supports trip, VMT, PMT and mode share calculations.')
        return
    if subregion.lower() not in ('boston','neighboring','i93','i495','region'):
        print('Only supports within boston, "neighboring" for towns neighboring Boston, I93, I495 or Region.')
        return
    

    subregion_dict = {'boston':'BOSTON','neighboring':'BOS_AND_NEI','i93':'in_i95i93','i495':'in_i495'}
    
            
    if metric.lower() == 'vmt':
        skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
        vmt_table = np.zeros((md.max_zone,md.max_zone))
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] / md.AO_dict[mode] for mode in md.auto_modes if mode in modes])
                        vmt_table += trip_table * skim_dict[peak]['Length (Skim)']
                        
        if subregion.lower() in subregion_dict:
            field = subregion_dict[subregion.lower()]
            #boston_o_auto_vmt = vmt_table[md.taz['BOSTON'],:][:, md.taz[field]== True]
            #boston_d_auto_vmt = vmt_table[md.taz[field]== True,:][:,md.taz['BOSTON']]
            boston_o_auto_vmt = mtx.OD_slice(vmt_table, O_slice = md.taz['BOSTON'], D_slice = md.taz[field]== True)
            boston_d_auto_vmt = mtx.OD_slice(vmt_table, O_slice = md.taz[field]== True, D_slice = md.taz['BOSTON'])
            
            town_definition = md.taz[md.taz[field]== True]
        
        elif subregion.lower() == 'region':
            # boston_o_auto_vmt = vmt_table[md.taz['BOSTON'],:]
            # boston_d_auto_vmt = vmt_table[:][:,md.taz['BOSTON']]
            boston_o_auto_vmt = mtx.OD_slice(vmt_table, O_slice = md.taz['BOSTON'])
            boston_d_auto_vmt = mtx.OD_slice(vmt_table, D_slice = md.taz['BOSTON'])
            
            town_definition = md.taz   
        
        zone_vmt_daily_o = pd.DataFrame(np.sum(boston_o_auto_vmt,axis=1)/2 ,columns=["VMT"])
        zone_vmt_daily_d = pd.DataFrame(np.sum(boston_d_auto_vmt,axis=0)/2 ,columns=["VMT"])

        town_vmt_o=pd.concat([town_definition,zone_vmt_daily_o],axis=1,join='inner')
        town_vmt_d=pd.concat([town_definition,zone_vmt_daily_d],axis=1,join='inner')

        vmt_sum_o = town_vmt_o[town_vmt_o['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['VMT']
        vmt_sum_d = town_vmt_d[town_vmt_d['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['VMT']

        subregion_vmt = (vmt_sum_o + vmt_sum_d).values[0]
        
        return subregion_vmt
    
    elif metric.lower() == 'pmt':
        skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
        
        
        pmt_table = np.zeros((md.max_zone,md.max_zone))
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        
                        trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                        for mode in md.modes if mode in drive_modes])
                        
                        pmt_table += trip_table * skim_dict[peak]['Length (Skim)']
        
        if subregion.lower() in subregion_dict:
            field = subregion_dict[subregion.lower()]
            boston_o_auto_pmt = mtx.OD_slice(pmt_table, O_slice = md.taz['BOSTON'],D_slice = md.taz[field]== True)
            boston_d_auto_pmt = mtx.OD_slice(pmt_table ,O_slice = md.taz[field]== True, D_slice = md.taz['BOSTON'])
            #boston_o_auto_pmt = pmt_table[md.taz['BOSTON'],:][:, md.taz[field]== True]
            #boston_d_auto_pmt = pmt_table[md.taz[field]== True,:][:,md.taz['BOSTON']]
            town_definition = md.taz[md.taz[field]== True]
        
        elif subregion.lower() == 'region':
            boston_o_auto_pmt = mtx.OD_slice(pmt_table, O_slice = md.taz['BOSTON'])
            boston_d_auto_pmt = mtx.OD_slice(pmt_table, D_slice = md.taz['BOSTON'])
            
            #boston_o_auto_pmt = pmt_table[md.taz['BOSTON'],:]
            #boston_d_auto_pmt = pmt_table[:][:,md.taz['BOSTON']]
            town_definition = md.taz
        
        zone_pmt_daily_o = pd.DataFrame(np.sum(boston_o_auto_pmt,axis=1)/2 ,columns=["PMT"])
        zone_pmt_daily_d = pd.DataFrame(np.sum(boston_d_auto_pmt,axis=0)/2 ,columns=["PMT"])

        town_pmt_o=pd.concat([town_definition,zone_pmt_daily_o],axis=1,join='inner')
        town_pmt_d=pd.concat([town_definition,zone_pmt_daily_d],axis=1,join='inner')

        pmt_sum_o = town_pmt_o[town_pmt_o['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['PMT']
        pmt_sum_d = town_pmt_d[town_pmt_d['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['PMT']

        boston_portion_pmt = (pmt_sum_o + pmt_sum_d).values[0]
        
        return boston_portion_pmt
    
    elif metric.lower() == 'pmt_act':
        skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
        
        
        pmt_table = np.zeros((md.max_zone,md.max_zone))
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        
                        trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                        for mode in ['Walk','Bike'] if mode in drive_modes])
                        
                        pmt_table += trip_table * skim_dict[peak]['Length (Skim)']
        
        if subregion.lower() in subregion_dict:
            field = subregion_dict[subregion.lower()]
            
            boston_o_auto_pmt = mtx.OD_slice(pmt_table, O_slice = md.taz['BOSTON'],D_slice = md.taz[field]== True)
            boston_d_auto_pmt = mtx.OD_slice(pmt_table, O_slice = md.taz[field]== True, D_slice = md.taz['BOSTON'])
            
            #boston_o_auto_pmt = pmt_table[md.taz['BOSTON'],:][:, md.taz[field]== True]
            #boston_d_auto_pmt = pmt_table[md.taz[field]== True,:][:,md.taz['BOSTON']]
            town_definition = md.taz[md.taz[field]== True]
        
        elif subregion.lower() == 'region':
            boston_o_auto_pmt = mtx.OD_slice(pmt_table,O_slice = md.taz['BOSTON'])
            boston_d_auto_pmt = mtx.OD_slice(pmt_table,D_slice = md.taz['BOSTON'])
            #boston_o_auto_pmt = pmt_table[md.taz['BOSTON'],:]
            #boston_d_auto_pmt = pmt_table[:][:,md.taz['BOSTON']]
            town_definition = md.taz
        
        zone_pmt_daily_o = pd.DataFrame(np.sum(boston_o_auto_pmt,axis=1)/2 ,columns=["PMT"])
        zone_pmt_daily_d = pd.DataFrame(np.sum(boston_d_auto_pmt,axis=0)/2 ,columns=["PMT"])

        town_pmt_o=pd.concat([town_definition,zone_pmt_daily_o],axis=1,join='inner')
        town_pmt_d=pd.concat([town_definition,zone_pmt_daily_d],axis=1,join='inner')

        pmt_sum_o = town_pmt_o[town_pmt_o['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['PMT']
        pmt_sum_d = town_pmt_d[town_pmt_d['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['PMT']

        boston_portion_pmt = (pmt_sum_o + pmt_sum_d).values[0]
        
        return boston_portion_pmt
    elif metric.lower() == 'mode share':        
        
        share_table = dict(zip(['drive','non-motorized','transit','smart mobility'],[0,0,0,0]))
        if subregion.lower() in subregion_dict:
            field = subregion_dict[subregion.lower()]
            for purpose in md.purposes:
                for peak in ['PK','OP']:
                    for veh_own in ['0','1']:
                        if mc_obj.table_container.get_table(purpose):
                            for mode in mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}']:
                                trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode]
                                boston_ii_trips = trip_table[md.taz['BOSTON'],:][:,md.taz['BOSTON']].sum()
                                trips = trip_table[md.taz['BOSTON'],:][:, md.taz[field]== True].sum() + trip_table[md.taz[field]== True,:][:,md.taz['BOSTON']].sum() - boston_ii_trips
                                category = md.mode_categories[mode]
                                share_table[category]+=trips
                                
        elif subregion.lower() == 'region':
            for purpose in md.purposes:
                for peak in ['PK','OP']:
                    for veh_own in ['0','1']:
                        if mc_obj.table_container.get_table(purpose):
                            for mode in mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}']:
                                trip_table = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode]
                                boston_ii_trips = trip_table[md.taz['BOSTON'],:][:,md.taz['BOSTON']].sum()
                                trips = trip_table[md.taz['BOSTON'],:][:].sum() + trip_table[:][:,md.taz['BOSTON']].sum() - boston_ii_trips
                                category = md.mode_categories[mode]
                                share_table[category]+=trips
        # normalize
        return (pd.DataFrame.from_dict(share_table,orient = 'index') / (pd.DataFrame.from_dict(share_table,orient = 'index').sum())).to_dict()[0]
    elif metric.lower() == 'trip':
        skim_dict = {'PK': mc_obj.drive_skim_PK,'OP':mc_obj.drive_skim_OP}
        tripsum_table = np.zeros((md.max_zone,md.max_zone))
        for purpose in md.purposes:
            for peak in ['PK','OP']:
                for veh_own in ['0','1']:
                    if mc_obj.table_container.get_table(purpose):
                        drive_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                        trip_table = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                        for mode in md.modes if mode in drive_modes])
                        tripsum_table += trip_table
                        
        if subregion.lower() in subregion_dict:
            field = subregion_dict[subregion.lower()]
            
            boston_o_trip = mtx.OD_slice(tripsum_table, O_slice = md.taz['BOSTON'],D_slice = md.taz[field]== True)
            boston_d_trip = mtx.OD_slice(tripsum_table, O_slice = md.taz[field]== True,D_slice = md.taz['BOSTON'])
            
            #boston_o_trip = tripsum_table[md.taz['BOSTON'],:][:, md.taz[field]== True]
            #boston_d_trip = tripsum_table[md.taz[field]== True,:][:,md.taz['BOSTON']]
            town_definition = md.taz[md.taz[field]== True]
        
        elif subregion.lower() == 'region':
            boston_o_trip = mtx.OD_slice(tripsum_table, O_slice = md.taz['BOSTON'])
            boston_d_trip = mtx.OD_slice(tripsum_table, D_slice = md.taz['BOSTON'])
            
            #boston_o_trip = tripsum_table[md.taz['BOSTON'],:]
            #boston_d_trip = tripsum_table[:][:,md.taz['BOSTON']]
            town_definition = md.taz   
        
        zone_daily_o = pd.DataFrame(np.sum(boston_o_trip,axis=1) ,columns=["trips"])
        zone_daily_d = pd.DataFrame(np.sum(boston_d_trip,axis=0) ,columns=["trips"])

        town_o=pd.concat([town_definition,zone_daily_o],axis=1,join='inner')
        town_d=pd.concat([town_definition,zone_daily_d],axis=1,join='inner')

        sum_o = town_o[town_o['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['trips']
        sum_d = town_d[town_d['TOWN']=='BOSTON,MA'].groupby(['TOWN']).sum()['trips']

        subregion_trip = (sum_o + sum_d).values[0]
        
        return subregion_trip

def __trips_to_from_boston(taz, mode, tripsum_table):
    boston_o_trip = mtx.OD_slice(tripsum_table, O_slice = taz['BOSTON'])
    boston_d_trip = mtx.OD_slice(tripsum_table, D_slice = taz['BOSTON'])
    #boston_o_trip = tripsum_table[taz['BOSTON'],:]
    #boston_d_trip = tripsum_table[:][:,taz['BOSTON']]
    zone_daily_o = pd.DataFrame(np.sum(boston_o_trip,axis=1) ,columns=[mode])
    zone_daily_d = pd.DataFrame(np.sum(boston_d_trip,axis=0) ,columns=[mode])
    town_o=pd.concat([taz,zone_daily_o],axis=1,join='inner')
    town_d=pd.concat([taz,zone_daily_d],axis=1,join='inner')

    zone_o = town_o[town_o['TOWN']=='BOSTON,MA'].groupby([md.taz_ID_field]).sum()[mode]
    zone_d = town_d[town_d['TOWN']=='BOSTON,MA'].groupby([md.taz_ID_field]).sum()[mode]
    return zone_o, zone_d

def trips_by_mode(mc_obj, mode='all'):

    auto_trip = np.zeros((md.max_zone,md.max_zone))
    transit_trip = np.zeros((md.max_zone,md.max_zone))
    nm_trip = np.zeros((md.max_zone,md.max_zone))
    sm_trip = np.zeros((md.max_zone,md.max_zone))
    
    for purpose in md.purposes:
        for peak in ['PK','OP']:
            for veh_own in ['0','1']:
                if mc_obj.table_container.get_table(purpose):
                    avail_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.drive_modes])
                    auto_trip += trips    

                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.transit_modes])
                    transit_trip += trips    

                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.active_modes])
                    nm_trip += trips    

                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.smart_mobility_modes])
                    sm_trip += trips                        
                    
    
    auto_o, auto_d = __trips_to_from_boston(md.taz, "auto", auto_trip)
    transit_o, transit_d = __trips_to_from_boston(md.taz, "transit", transit_trip)
    nm_o, nm_d = __trips_to_from_boston(md.taz, "nm", nm_trip)
    sm_o, sm_d = __trips_to_from_boston(md.taz, "sm", sm_trip)
    
    trips_o = auto_o.to_frame().join(transit_o)
    trips_o = trips_o.join(nm_o)
    trips_o = trips_o.join(sm_o)

    trips_d = auto_d.to_frame().join(transit_d)
    trips_d = trips_d.join(nm_d)
    trips_d = trips_d.join(sm_d)
    
    trips_o.to_csv(mc_obj.config.out_path + 'trip_p_mode_zone.csv')
    trips_d.to_csv(mc_obj.config.out_path + 'trip_a_mode_zone.csv')

def __trips_to_region(mask, taz, mode, tripsum_table): 
    
    boston_o_sums = np.sum(mtx.OD_slice(tripsum_table, O_slice = taz['BOSTON'],D_slice = mask==1),axis=1)
    nonboston_o_sums = np.sum(mtx.OD_slice(tripsum_table,D_slice = ((mask * (taz['BOSTON']).values)==1)),axis=1)
    
    boston_zone = taz.join(pd.DataFrame(boston_o_sums,columns=[mode + "boston"]))
    nonboston_zone = taz.join(pd.DataFrame(nonboston_o_sums,columns=[mode + "nonboston"]))
    nonboston_zone.loc[nonboston_zone['TOWN']=='BOSTON,MA',mode + 'nonboston']=0   
    
    zone_daily = pd.DataFrame(boston_zone[mode + "boston"]).join(nonboston_zone[mode + 'nonboston'])
    
    return pd.DataFrame(zone_daily.sum(axis=1), columns=[mode])


def productions_by_region(mc_obj, region='all', cordon_area=[]):
    
    
    auto_trip = np.zeros((md.max_zone,md.max_zone))
    da_trip = np.zeros((md.max_zone,md.max_zone))
    sr_trip = np.zeros((md.max_zone,md.max_zone))  
    wat_trip = np.zeros((md.max_zone,md.max_zone))
    dat_trip = np.zeros((md.max_zone,md.max_zone))
    nm_trip = np.zeros((md.max_zone,md.max_zone))
    smra_trip = np.zeros((md.max_zone,md.max_zone))
    smsh_trip = np.zeros((md.max_zone,md.max_zone))
    
    for purpose in md.purposes:
        for peak in ['PK','OP']:
            for veh_own in ['0','1']:
                if mc_obj.table_container.get_table(purpose):
                    avail_modes = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'].keys()
                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.drive_modes])
                    auto_trip += trips    
                    
                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.da_mode])
                    da_trip += trips    

                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.sr_mode])
                    sr_trip += trips                        

                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.WAT_modes])
                    wat_trip += trips    

                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.DAT_modes])
                    dat_trip += trips    

                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.active_modes])
                    nm_trip += trips    

                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.sm_ride_alone])
                    smra_trip += trips                        

                    trips = sum([mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode] 
                                    for mode in avail_modes if mode in md.sm_shared_ride])
                    smsh_trip += trips                       
    
    if region=='all':
        mask = np.ones(md.max_zone)
        outfile = 'trip_p_to_region.csv'
    elif region=='Boston':
        mask = (taz['BOSTON']).values * 1 #[1]*447 + [0]*(md.max_zone - 447)
        outfile = 'trip_p_to_boston.csv'       
    elif region=='cordon':
        mask = taz['BOSTON_NB'].isin(cordon_area).values * 1
        outfile = 'trip_p_to_cordon.csv'           
    
       
    auto_d = __trips_to_region(mask, md.taz, "auto", auto_trip)
    da_d = __trips_to_region(mask, md.taz, "da", da_trip)
    sr_d = __trips_to_region(mask, md.taz, "sr", sr_trip)
    wat_d = __trips_to_region(mask, md.taz, "wat", wat_trip)
    dat_d = __trips_to_region(mask, md.taz, "dat", dat_trip)
    nm_d = __trips_to_region(mask, md.taz, "nm", nm_trip)
    smra_d = __trips_to_region(mask, md.taz, "smra", smra_trip)
    smsh_d = __trips_to_region(mask, md.taz, "smsh", smsh_trip)

    trips_d = auto_d.join(da_d)
    trips_d = trips_d.join(sr_d)
    trips_d = trips_d.join(wat_d)
    trips_d = trips_d.join(dat_d)
    trips_d = trips_d.join(nm_d)
    trips_d = trips_d.join(smra_d)
    trips_d = trips_d.join(smsh_d)
    trips_d = taz.join(trips_d)
    
    trips_d.to_csv(mc_obj.config.out_path + outfile)

        
def write_summary_by_subregion(mc_obj, by='all'):

    '''
    Summarizes VMT, PMT or mode share by subregions of Massachusetts surrounding Boston (neighboring towns of Boston / within I-93/95 / within I-495).
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param taz_fn: TAZ file that contains subregion definition
    :param out_path: output path.
    '''
    subregion_dict = dict(zip(['boston','neighboring','i93','i495','region'],['Within Boston','Boston and Neighboring Towns', 'Within I-93/95', 'Within I-495', 'Entire Region']))
    vmt_summary_df = pd.DataFrame(index = subregion_dict.values(), columns = ['VMT to/from Boston'])
    pmt_summary_df = pd.DataFrame(index = subregion_dict.values(), columns = ['PMT to/from Boston'])
    pmtact_summary_df = pd.DataFrame(index = subregion_dict.values(), columns = ['PMT Active Modes to/from Boston'])
    mode_share_df = pd.DataFrame(index = subregion_dict.values(),columns = ['drive','non-motorized','transit','smart mobility'])
    trip_summary_df = pd.DataFrame(index = subregion_dict.values(),columns = ['Trips to/from Boston'])
    for subregion in subregion_dict:
        vmt_summary_df.loc[subregion_dict[subregion]] = __compute_summary_by_subregion(mc_obj, metric = 'VMT',subregion = subregion)
        pmt_summary_df.loc[subregion_dict[subregion]] = __compute_summary_by_subregion(mc_obj, metric = 'PMT',subregion = subregion)
        pmtact_summary_df.loc[subregion_dict[subregion]] = __compute_summary_by_subregion(mc_obj, metric = 'pmt_act',subregion = subregion)
        mode_share_df.loc[subregion_dict[subregion]] = __compute_summary_by_subregion(mc_obj, metric = 'mode share',subregion = subregion)
        trip_summary_df.loc[subregion_dict[subregion]] = __compute_summary_by_subregion(mc_obj, metric = 'trip',subregion = subregion)
    vmt_summary_df.to_csv(mc_obj.config.out_path + 'vmt_summary_subregions.csv')
    pmt_summary_df.to_csv(mc_obj.config.out_path + 'pmt_summary_subregions.csv')
    pmtact_summary_df.to_csv(mc_obj.config.out_path + 'act_pmt_summary_subregions.csv')
    mode_share_df.to_csv(mc_obj.config.out_path + 'mode_share_summary_subregions.csv')
    trip_summary_df.to_csv(mc_obj.config.out_path + 'trip_summary_subregions.csv')
    
def write_summary_by_subregion_sm(mc_obj, by='all'):

    '''
    Summarizes Smart Mobility VMT and trips by subregions of Massachusetts surrounding Boston (neighboring towns of Boston / within I-93/95 / within I-495).
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param taz_fn: TAZ file that contains subregion definition
    :param out_path: output path.
    '''
    subregion_dict = dict(zip(['boston','neighboring','i93','i495','region'],['Within Boston','Boston and Neighboring Towns', 'Within I-93/95', 'Within I-495', 'Entire Region']))    
    smra_vmt_summary_df = pd.DataFrame(index = subregion_dict.values(), columns = ['VMT to/from Boston'])
    smsh_vmt_summary_df = pd.DataFrame(index = subregion_dict.values(), columns = ['VMT to/from Boston'])
    smra_trip_summary_df = pd.DataFrame(index = subregion_dict.values(),columns = ['Trips to/from Boston'])
    smsh_trip_summary_df = pd.DataFrame(index = subregion_dict.values(),columns = ['Trips to/from Boston'])
    for subregion in subregion_dict:
        smra_vmt_summary_df.loc[subregion_dict[subregion]] = __sm_compute_summary_by_subregion(mc_obj, metric = 'VMT',subregion = subregion, sm_mode='SM_RA')
        smsh_vmt_summary_df.loc[subregion_dict[subregion]] = __sm_compute_summary_by_subregion(mc_obj, metric = 'VMT',subregion = subregion, sm_mode='SM_SH')
        smra_trip_summary_df.loc[subregion_dict[subregion]] = __sm_compute_summary_by_subregion(mc_obj, metric = 'trip',subregion = subregion, sm_mode='SM_RA')
        smsh_trip_summary_df.loc[subregion_dict[subregion]] = __sm_compute_summary_by_subregion(mc_obj, metric = 'trip',subregion = subregion, sm_mode='SM_SH')
    smra_vmt_summary_df.to_csv(mc_obj.config.out_path + 'sm_ra_vmt_summary_subregions.csv')
    smsh_vmt_summary_df.to_csv(mc_obj.config.out_path + 'sm_sh_vmt_summary_subregions.csv')
    smra_trip_summary_df.to_csv(mc_obj.config.out_path + 'sm_ra_trip_summary_subregions.csv')
    smsh_trip_summary_df.to_csv(mc_obj.config.out_path + 'sm_sh_trip_summary_subregions.csv')
        
def transit_ridership(mc_obj, by='all'):
    '''
    Summarizes transit ridership by peak period in cities and towns with MBTA subway service.
    
    :param mc_obj: mode choice module object as defined in the IPython notebook
    :param mbta_fn: TAZ file that contains MBTA coverage definition
    :param out_path: output path.
    '''
    MBTA_fn =mc_obj.config.data_path + "..\MBTA_coverage.csv"
    MBTA_cvg = pd.read_csv(MBTA_fn)
    taz_cvg = mc_obj.taz.merge(MBTA_cvg, how = 'left', on = 'TOWN')
    taz_cvg = taz_cvg[['ID_FOR_CS','subway','TOWN']]
    taz_cvg['covered'] = taz_cvg['subway']==1 # 870 TAZs included.
    ridership_master = pd.DataFrame(columns=['region','subway'])
    for purpose in md.purposes:
        for peak in ['PK','OP']:
            for veh_own in ['0','1']:
                ridership = pd.DataFrame(index=range(0,2),columns=['region','subway']).fillna(0)
                if mc_obj.table_container.get_table(purpose):
                    for mode in set(mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'])&set(md.transit_modes):
                        boston_ii = mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode][(taz_cvg['TOWN']=='BOSTON,MA'),:][:,(taz_cvg['TOWN']=='BOSTON,MA')].sum()  
                        ridership['subway'] += (mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode][taz_cvg['covered'],:][:,(taz_cvg['TOWN']=='BOSTON,MA')].sum() +  
                                                mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode][(taz_cvg['TOWN']=='BOSTON,MA'),:][:,taz_cvg['covered']].sum() - 
                                                boston_ii)
                        ridership['region'] += (mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode][:][:,(taz_cvg['TOWN']=='BOSTON,MA')].sum() + 
                                                mc_obj.table_container.get_table(purpose)[f'{veh_own}_{peak}'][mode][(taz_cvg['TOWN']=='BOSTON,MA'),:].sum()  - boston_ii)
                        ridership['peak'] = peak
                    ridership_master = ridership_master.append(ridership.reset_index(), sort = True)
    #ridership_summary = ridership_master.groupby(['peak']).sum()
    # calculate ridership
    ridership_master.groupby('peak').sum().to_csv(mc_obj.config.out_path + 'transit_ridership_summary.csv')
    